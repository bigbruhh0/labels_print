import re
with open('data/brand_names.txt', 'r') as file:
    lines = file.readlines()
struct_len=[]
for i in range(100):
	struct_len.append([])
for i in lines:
	l=len(i)
	struct_len[l].append(i)
for i in range(len(struct_len)):
	if len(struct_len[i])>0:
		print(i,len(struct_len[i]),struct_len[i][0])

import openpyxl
file_path="data/data_big.xlsx"
def read_excel_column(file_path, column_index):
    # Открыть файл Excel
    wb = openpyxl.load_workbook(file_path)
    # Получить активный лист
    sheet = wb.active
    # Считать данные из указанного столбца в список
    column_data = [sheet.cell(row=row_index, column=column_index).value for row_index in range(1, sheet.max_row + 1)]
    return column_data

def write_to_txt(file_path, data):
    with open(file_path, 'w', encoding='utf-8') as f:
        for item in data:
            f.write("%s\n" % item)

def main():
	bems=[]
	# Список1
	search_list1 = ["EAU DE "]
	# Список2
	search_list2 = ["EAU DE PARFUM", "EAU DE COLOGNE", "ELIXIR DE PARFUM", "POUR HOMME","POUR FEMME","POUR ELLE"]
	# Путь к файлу Excel
	excel_file_path = file_path
	# Индекс столбца, который нужно считать
	column_index = 2

	# Чтение данных из Excel-файла
	column_data = read_excel_column(excel_file_path, column_index)
	
	filtered_data = []
	for item in column_data:
		if item is not None:
			
			a=str(item).upper()
			if any(keyword in a for keyword in search_list1) and not any(keyword in a for keyword in search_list2):
				pattern = r'\bEAU DE (\w+(?: \w+)*)\b'
				bux=re.search(pattern, a)
				if bux:
					filtered_data.append("EAU DE "+bux.group(1))
				
				
	filtered_data=set(filtered_data)
	# Запись результатов в файл txt_res.txt
	write_to_txt("data/txt_res.txt", filtered_data)

if __name__ == "__main__":
    main()
